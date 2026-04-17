from typing import Any

from openpyxl import load_workbook

from app.extractors.base import Extractor


class XlsxExtractor(Extractor):
    file_type = "xlsx"

    def extract_elements(self, path: str) -> list[dict[str, Any]]:
        wb = load_workbook(path, data_only=True, read_only=True)
        elements: list[dict[str, Any]] = []
        try:
            for ws in wb.worksheets:
                rows: list[list[str]] = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([_cell_str(v) for v in row])
                elements.append({"type": "sheet", "name": ws.title, "rows": rows})
        finally:
            wb.close()
        return elements


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v)
